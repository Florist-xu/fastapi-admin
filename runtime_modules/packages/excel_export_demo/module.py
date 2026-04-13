from __future__ import annotations

import re
from io import BytesIO
from typing import Any, cast
from urllib.parse import quote

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from runtime_modules.base import RuntimeModuleBase, RuntimeModuleRoute


INVALID_FILENAME_RE = re.compile(r'[\\/:*?"<>|]+')
INVALID_SHEET_RE = re.compile(r'[:\\/?*\[\]]+')


class Module(RuntimeModuleBase):
    def get_routes(self) -> list[RuntimeModuleRoute]:
        return [
            RuntimeModuleRoute(path="/meta", methods=["GET"], handler="meta", summary="Module meta"),
            RuntimeModuleRoute(path="/export", methods=["POST"], handler="export_file", summary="Export xlsx"),
        ]

    async def meta(self, request) -> dict[str, Any]:
        return {
            "title": self.context.name,
            "version": self.context.version,
            "usage": {
                "endpoint": f"/runtime-module/execute/{self.context.code}/export",
                "method": "POST",
                "body": {
                    "filename": "articles",
                    "sheet_name": "Articles",
                    "columns": [
                        {"key": "title", "title": "Title", "width": 28},
                        {"key": "status", "title": "Status", "width": 14},
                    ],
                    "rows": [
                        {"title": "AI overview", "status": "Published"},
                    ],
                },
            },
            "config": self.context.config,
        }

    def _sanitize_filename(self, value: str | None) -> str:
        raw = str(value or "export").strip()
        safe = INVALID_FILENAME_RE.sub("_", raw).strip(" ._")
        return safe[:80] or "export"

    def _sanitize_sheet_name(self, value: str | None) -> str:
        raw = str(value or self.context.config.get("default_sheet_name") or "Sheet1").strip()
        safe = INVALID_SHEET_RE.sub("_", raw).strip("'")
        return (safe[:31] or "Sheet1").strip() or "Sheet1"

    def _normalize_columns(self, payload: dict[str, Any]) -> list[dict[str, Any]]:
        raw_columns = payload.get("columns") or []
        rows = payload.get("rows") or []

        if raw_columns:
            columns: list[dict[str, Any]] = []
            for item in raw_columns:
                key = str(item.get("key") or "").strip()
                title = str(item.get("title") or "").strip()
                if not key or not title:
                    continue
                width = item.get("width")
                columns.append(
                    {
                        "key": key,
                        "title": title,
                        "width": max(10, min(int(width), 60)) if width is not None else None,
                    }
                )
            if columns:
                return columns

        if isinstance(rows, list) and rows:
            first_row = rows[0] if isinstance(rows[0], dict) else {}
            return [{"key": key, "title": key, "width": None} for key in first_row.keys()]

        raise ValueError("columns cannot be empty")

    def _normalize_rows(self, payload: dict[str, Any]) -> list[dict[str, Any]]:
        rows = payload.get("rows")
        if not isinstance(rows, list) or not rows:
            raise ValueError("rows cannot be empty")

        normalized_rows: list[dict[str, Any]] = []
        for item in rows:
            if isinstance(item, dict):
                normalized_rows.append(item)

        if not normalized_rows:
            raise ValueError("rows must be a list of objects")
        return normalized_rows

    def _build_workbook(
        self,
        *,
        columns: list[dict[str, Any]],
        rows: list[dict[str, Any]],
        sheet_name: str,
    ) -> bytes:
        workbook = Workbook()
        sheet = cast(Worksheet, workbook.active)
        sheet.title = sheet_name
        sheet.freeze_panes = "A2"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
        header_alignment = Alignment(vertical="center", horizontal="center")
        body_alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

        headers = [str(item["title"]) for item in columns]
        sheet.append(headers)

        for col_index, header in enumerate(headers, start=1):
            header_cell = cast(Cell, sheet.cell(row=1, column=col_index, value=header))
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = header_alignment

        for row in rows:
            values: list[Any] = []
            for item in columns:
                value = row.get(item["key"])
                if isinstance(value, list):
                    value = ", ".join(str(entry) for entry in value if entry is not None)
                elif value is None:
                    value = ""
                values.append(value)
            sheet.append(values)

        for row_cells in sheet.iter_rows(min_row=2):
            for cell in row_cells:
                cast(Cell, cell).alignment = body_alignment

        for index, item in enumerate(columns, start=1):
            width = item.get("width")
            if width is None:
                sample_lengths = [len(str(item["title"]))]
                sample_lengths.extend(len(str(row.get(item["key"], ""))) for row in rows[:100])
                width = min(max(max(sample_lengths, default=12) + 2, 12), 48)
            sheet.column_dimensions[get_column_letter(index)].width = float(width)

        if self.context.config.get("auto_filter", True):
            sheet.auto_filter.ref = sheet.dimensions

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def export_file(self, request):
        payload = await request.json()
        columns = self._normalize_columns(payload)
        rows = self._normalize_rows(payload)

        filename = self._sanitize_filename(payload.get("filename"))
        sheet_name = self._sanitize_sheet_name(payload.get("sheet_name"))
        content = self._build_workbook(columns=columns, rows=rows, sheet_name=sheet_name)

        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
